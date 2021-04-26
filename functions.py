from selenium import webdriver
import sqlite3
import mapbox
from selenium.webdriver.common.keys import Keys
import os
import json
import datetime
import time
import copy
import pymongo
import logging
import openpyxl
from openpyxl.styles import PatternFill, Fill, Color
from openpyxl.styles import colors
basic_url = 'https://jnportal.ujn.gov.rs/'
from za_listu import cpv_dict
log = open('log.txt','w')


def distanca(mesto):
    destination = str(mesto).replace('đ','dj') + ', Serbia'
    db = sqlite3.connect("placesDB.db")
    map = mapbox.Geocoder(access_token="pk.eyJ1IjoiZHVoaXpqYW1lIiwiYSI6ImNrZHhzdGxvbDM0aGkyd21xNHVleTByajUifQ.esMht2IXyPcuICHLA9vl2Q")
    d_code = map.forward(destination)
    dest = d_code.json()
    y = 0
    nasao = False
    while(dest['features'][int(y)]['place_type']!=["place"] or dest['features'][int(y)]['place_type']!=["region"]  or dest['features'][int(y)]['text']=="Serbia"):
        for i,z in enumerate(dest['features']):
            if (z['place_type']==['place'] or z['place_type']==['region']) and z['text']!="Serbia":
                # print(z)
                y = i
                nasao = True
                break
        if nasao == True:
            break
        x = str(mesto).replace('đ','dj').replace('š','s').replace('ž','z').replace('č','c').replace('ć','c').replace('Đ','Dj').replace('Š','S').replace('Ž','Z').replace('Č','C').replace('Ć','C')
        execution = f"SELECT * from mesto where nazivMesta = \"{x}\""
        cursor = db.execute(execution)
        res = cursor.fetchone()
        if res is None:
            print("Can't find place in database.")
            return
        zipCode = res[0]
        while(zipCode % 100 !=0 ):
            zipCode = zipCode - 1
            cursor = db.execute(f"SELECT * from mesto where idMesta = {zipCode}")
            mesto = cursor.fetchone()
            if mesto is None:
                continue
            d_code = map.forward(str(mesto) + ', Serbia')
            dest = d_code.json()
            print(dest['features'][y]['text'])
            if (dest['features'][y]['place_type']==["place"] or dest['features'][int(y)]['place_type']==["region"]) and dest['features'][y]['text']!="Serbia":
                break
    beograd = map.forward("Belgrade, Serbia")
    novi_sad = map.forward("Novi Sad, Serbia")
    lazarevac = map.forward("Lazarevac, Serbia")

    bg_json = beograd.json()
    la_json = lazarevac.json()
    ns_json = novi_sad.json()

    origins = []
    origins.append((bg_json['features'][0]['text'],bg_json['features'][0]['center']))
    origins.append((la_json['features'][0]['text'],la_json['features'][0]['center']))
    origins.append((ns_json['features'][0]['text'],ns_json['features'][0]['center']))
    destinations = {
        'type': 'Feature',
        'properties': {'name': dest['features'][y]['text']},
        'geometry': {
            'type': 'Point',
            'coordinates': dest['features'][y]['center']
            }
        }
    service = mapbox.Directions(access_token="pk.eyJ1IjoiZHVoaXpqYW1lIiwiYSI6ImNrZHhzdGxvbDM0aGkyd21xNHVleTByajUifQ.esMht2IXyPcuICHLA9vl2Q")
    for place in origins:
        origin = {
                'type': 'Feature',
                'properties': {'name': place[0]},
                'geometry': {
                    'type': 'Point',
                    'coordinates': place[1]
                    }
                }
        directions = service.directions([origin,destinations])
        dir_json = directions.json()
        distance = dir_json['routes'][0]['legs'][0]['distance']/1000

        if int(distance) <= 165 and (str(place[0]).lower() == 'Belgrade'.lower() or str(place[0]).lower() == "Beograd".lower()):
            return True
        elif int(distance) <= 60 and str(place[0]).lower() == 'Novi Sad'.lower():
            return True
        elif int(distance) <= 80 and str(place[0]).lower() == 'Lazarevac'.lower():
            return True
    return False


def delete_doc(name):
    namex = ""
    if '/' in str(name):
        buff = str(name).split('/')
        for part in buff:
            namex = namex + part
    else:
        namex = str(name)
    if(len(namex)>100):
        namex = namex[:100]
    else:
        namex = namex
    os.remove(namex + ".zip")

def money_output(money):
    string = money[::-1]
    if '.' in string:
        string = string.split('.')[1]
    output = ""
    for i in range(0,len(string)+1):
        if (i-1)%3 == 0:
            output = output + '.' + string[i-1:i+2]
    final = output[::-1]
    final = final[0:len(final)-1]
    return final



def get_doc(browser,id,name):

    url = basic_url + '/tender-eo/' + str(id)
    browser.get(url)
    time.sleep(3)

    place = browser.find_element_by_id("uiMainContentPH_uiMainContentPH_ucBasicDetails_uiNuts")
    mesto = str(place.text)
    word = mesto.split(',')
    if distanca(word[1]) == False:
        print(word[1])
        return False, False

    kod_elem = browser.find_element_by_css_selector('#tenderNoticesContainer > div:nth-child(1) > div:nth-child(6) > div:nth-child(1) > table:nth-child(1) > tbody:nth-child(2) > tr:nth-child(1) > td:nth-child(3)')
    kod = str(kod_elem.text)
    print(kod)

    elem = browser.find_element_by_css_selector('html body.skin-blue.sidebar-mini form#form1 div.wrapper div#uiContentWrapper.content-wrapper section#uiContent.content div#uiMainContentPH_uiTenderWrapper.notsubtracted div#documentControl_uiMainContentPH_uiMainContentPH_uiDocumentControl.dx-accordion.dx-widget.dx-visibility-change-handler.dx-collection div.dx-accordion-wrapper div.dx-item.dx-accordion-item.dx-item-selected.dx-accordion-item-opened div.dx-item-content.dx-accordion-item-title a')
    if elem is None:
        return
    elem.click()

    namex = ""
    if '/' in str(name):
        buff = str(name).split('/')
        for part in buff:
            namex = namex + part
    else:
        namex = str(name)
    if(len(namex)>100):
        namex = namex[:100]
    else:
        namex = namex
    time.sleep(3)
    while True:
        time.sleep(3)
        try:
            if os.path.getsize('konkursne/Konkursna dokumentacija za ponuđača.zip') > 0:
                os.rename("konkursne/Konkursna dokumentacija za ponuđača.zip", "konkursne/" + str(namex) + ".zip")
                break
        except:
            print("Nema fajla jos: " + namex)

    return word[1].strip(), kod




def set_browser_pref():
    profile = webdriver.FirefoxProfile()
    profile.set_preference("browser.download.dir",'/home/cole/Documents/scrapper/python/konkursne')
    profile.set_preference("browser.download.folderList",2)
    profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/plain,text/x-csv,text/csv,application/vnd.ms-excel,application/csv,application/x-csv,text/csv,text/comma-separated-values,text/x-comma-separated-values,text/tab-separated-values,application/pdf,application/zip")
    profile.set_preference("browser.download.manager.showWhenStarting",False)
    profile.set_preference("browser.helperApps.neverAsk.openFile","text/plain,text/x-csv,text/csv,application/vnd.ms-excel,application/csv,application/x-csv,text/csv,text/comma-separated-values,text/x-comma-separated-values,text/tab-separated-values,application/pdf,application/zip")
    profile.set_preference("browser.helperApps.alwaysAsk.force", False)
    profile.set_preference("browser.download.manager.useWindow", False)
    profile.set_preference("browser.download.manager.focusWhenStarting", False)
    profile.set_preference("browser.helperApps.neverAsk.openFile", "")
    profile.set_preference("browser.download.manager.alertOnEXEOpen", False)
    profile.set_preference("browser.download.manager.showAlertOnComplete", False)
    profile.set_preference("browser.download.manager.closeWhenDone", True)
    profile.set_preference("pdfjs.disabled", True)
    return profile

def login(username,password,browser):
    url = basic_url
    browser.get(url + 'prijava')
    username_button = browser.find_element_by_id('uiUsername')
    username_button.send_keys(username)
    pass_field = browser.find_element_by_id('uiPassword')
    pass_field.send_keys(password)
    log_button = browser.find_element_by_id('uiLogin')
    log_button.click()


def get_json_list(browser,date):
    params = f'postupci-svi?initFilter=["NoticePublishDate",">=","{date}"]'
    browser.get(basic_url + params)
    json_dl = browser.find_element_by_css_selector('html body.skin-blue.sidebar-mini form#form1 div.wrapper div#uiContentWrapper.content-wrapper section#uiContent.content div#searchGridContainer.dx-widget.searchGridContainer.dx-visibility-change-handler div.dx-datagrid.dx-gridbase-container div.dx-datagrid-header-panel div.dx-toolbar.dx-widget.dx-visibility-change-handler.dx-collection div.dx-toolbar-items-container div.dx-toolbar-after div.dx-item.dx-toolbar-item.dx-toolbar-button div.dx-item-content.dx-toolbar-item-content div.dx-button.dx-button-normal.dx-button-mode-contained.dx-widget.dx-button-has-icon div.dx-button-content i.dx-icon.fa.fa-file-o')
    curr_time = datetime.datetime.now()
    json_dl.click()
    times = str(curr_time.year) + '-' + str(curr_time.month-1) + '-' + str(curr_time.day) + '-' + str(curr_time.hour) + '-' + str(curr_time.minute) + '-' + str(curr_time.second)
    filename = 'konkursne/Postupci javnih nabavki_' + times + '.json'
    log.write(str(curr_time.month) + ' ' + str(curr_time.day))

    #TODO filename u while(true) petlji kao za get_doc
    while True:
        time.sleep(1)
        try:
            with open(filename,'r') as f:
                data = json.load(f)
                break
        except:
            time_d = datetime.timedelta(seconds=1)
            curr_time = curr_time + time_d
            times = str(curr_time.year) + '-' + str(curr_time.month-1) + '-' + str(curr_time.day) + '-' + str(curr_time.hour) + '-' + str(curr_time.minute) + '-' + str(curr_time.second)
            filename = 'konkursne/Postupci javnih nabavki_' + times + '.json'
            print(filename)

    try:
        os.remove(filename)
    except:
        print('Failed removal')
    return data

def enter_data(browser, data):
    wb = openpyxl.load_workbook("tabela.xlsx")
    sheet = wb["bELI"]

    counter = int(sheet['J2'].value)
    for i in data:
        id = i['Id']
        name = i['Name']
        value = i['EstimatedValue']
        tip = i['TypeContract']
        procId = i['ProcedureTypeId']
        cpv = i['CPVExtended'][:3]
        print(name)
        print(counter)
        log.write(name + '\n')
        if cpv not in cpv_dict:
            log.write('id = ' + str(id) + '\n')
            print('preskoceno zbog cpv')
            log.write('preskoceno zbog cpv' + '\n')
            # log.write(str(cpv_dict[cpv]) + '\n')
            log.write('=============================================' + '\n')
            continue
        print(cpv_dict[cpv])
        print('id = ' + str(id))
        deadline = i['SubmissionDeadline']
        if deadline:
            rok = ' '.join(deadline.split('T'))
        else:
            rok = 'Nema roka za podnosenje'
        narucilac = i['BusinessEntityName']
        if procId !=1:
            log.write('id = ' + str(id) + '\n')
            print('preskoceno neotovrenog ugovora\n')
            log.write('preskoceno zbog neotvorenog ugovora\n')
            log.write('=============================================' + '\n')
            continue
        mesto, kod = get_doc(browser,id,name)
        if not mesto:
            print('preskoceno zbog mesta')
            log.write('id = ' + str(id) + '\n')
            log.write('preskoceno zbog mesta\n')
            log.write(str(mesto) + '\n')
            log.write('=============================================' + '\n')
            # time.sleep(1)
            continue
        else:
            log.write('id = ' + str(id) + '\n')
            sheet['A'+str(counter)].value = str(name)
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            sheet['A'+str(counter)].fill = my_fill
            sheet['B'+str(counter)].value = str(kod)
            sheet['C'+str(counter)].value = str(narucilac)
            sheet['D'+str(counter)].value = str(mesto)
            sheet['E'+str(counter)].value = str(rok)
            sheet['F'+str(counter)].value = str(tip)
            if value == 'null' or value == 'None' or value is None:
                    sheet['G'+str(counter)].value = "Нема процењене вредности."
            else:
                if ',' not in str(value):
                    sheet['G'+str(counter)].value = "Процењена вредност: " + money_output(str(value)) + ",00 дин"
                else:
                    sheet['G'+str(counter)].value = "Процењена вредност: " + money_output(str(value)) + " дин"
            counter = counter + 1
            log.write('=============================================' + '\n')

    sheet['J2'] = counter
    wb.save("tabela.xlsx")
    print('saved to tabela.xlsx')
